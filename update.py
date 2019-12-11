from openpyxl import load_workbook
import sqlite3


def update(conn, cur, table_name):
    wb = load_workbook(filename=table_name + '.xlsx', read_only=True)
    if table_name == '销售人员业务跟踪表':
        ws = wb['page']
    else:
        ws = wb['page1']

    max_col = ws.max_column
    max_row = ws.max_row

    title_row = ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=max_col)

    titles = []

    for title_tuple in title_row:
        for key in title_tuple:
            titles.append(key.value)

    if table_name == '后线':
        titles.remove('同城异地')
        titles.append('备注')

    str_sql = f"CREATE TABLE '{table_name}' ("
    for value in titles:
        str_sql += f"'{value}' TEXT ,"
    str_sql = str_sql[:-2] + ")"

    cur.execute(str_sql)

    if table_name == '销售人员业务跟踪表':
        begin_row = 2
    else:
        begin_row = 3

    for row in ws.iter_rows(min_row=begin_row, max_row=max_row, min_col=1,
                            max_col=max_col):
        str_sql = f"INSERT INTO '{table_name}' VALUES ("
        for key in row:
            str_sql += f"'{key.value}', "
        str_sql = str_sql[:-2] + ')'
        cur.execute(str_sql)

    conn.commit()
