from openpyxl import load_workbook
import sqlite3

wb = load_workbook(filename='后线.xlsx',read_only=True)
ws = wb['page1']

max_col = ws.max_column
max_row = ws.max_row

title_row = ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=max_col)

titles = []

for title_tuple in title_row:
    for key in title_tuple:
        titles.append(key.value)

titles.remove('同城异地')

conn = sqlite3.connect('data.db')
cur = conn.cursor()

str_sql = "CREATE TABLE '后线' ("
for value in titles:
    str_sql += f"'{value}' TEXT ,"
str_sql = str_sql[:-2] + ")"

cur.execute(str_sql)

for row in ws.iter_rows(min_row=3, max_row=max_row, min_col=1, max_col=max_col-1):
    str_sql = "INSERT INTO '后线' VALUES ("
    for key in row:
        str_sql += f"'{key.value}', "
    str_sql = str_sql[:-2] + ')'
    print(str_sql)
    cur.execute(str_sql)

conn.commit()