import sqlite3

from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Color, Font, NamedStyle,
                             PatternFill, Side)

from write import write, write_zhong_zhi
from rens import rens_list
from update import update


conn = sqlite3.connect(':memory:')
cur = conn.cursor()

update(conn, cur, '后线')
update(conn, cur, '前线')
update(conn, cur, '销售人员业务跟踪表')

wb = Workbook()

title_style = NamedStyle(name='title_style')
title_style.font = Font(name='微软雅黑', size=12, bold=True)
title_style.alignment = Alignment(horizontal='center',
                                  vertical='center',
                                  wrapText=True)
title_style.border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

num_style = NamedStyle(name='num_style')
num_style.font = Font(name='微软雅黑', size=11)
num_style.alignment = Alignment(horizontal='center', vertical='center')
num_style.border = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))
num_style.number_format = '0.00'

num_style_fill = NamedStyle(name='num_style_fill')
num_style_fill.font = Font(name='微软雅黑', size=11)
num_style_fill.alignment = Alignment(horizontal='center', vertical='center')
num_style_fill.border = Border(left=Side(style='thin'),
                               right=Side(style='thin'),
                               top=Side(style='thin'),
                               bottom=Side(style='thin'))
num_style_fill.fill = PatternFill('solid', fgColor='E5E5E5')
num_style_fill.number_format = '0.00'

str_style = NamedStyle(name='str_style')
str_style.font = Font(name='微软雅黑', size=11)
str_style.alignment = Alignment(horizontal='center', vertical='center')
str_style.border = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))

str_style_fill = NamedStyle(name='str_style_fill')
str_style_fill.font = Font(name='微软雅黑', size=11)
str_style_fill.alignment = Alignment(horizontal='center', vertical='center')
str_style_fill.border = Border(left=Side(style='thin'),
                               right=Side(style='thin'),
                               top=Side(style='thin'),
                               bottom=Side(style='thin'))
str_style_fill.fill = PatternFill('solid', fgColor='E5E5E5')


wb.add_named_style(title_style)
wb.add_named_style(num_style)
wb.add_named_style(num_style_fill)
wb.add_named_style(str_style)
wb.add_named_style(str_style_fill)

rens = rens_list(cur, '前线')

write(wb, rens, '整体')
write(wb, rens, '车险')
write(wb, rens, '非车险')

write_zhong_zhi(wb, cur, rens, '分公司营业一部')
write_zhong_zhi(wb, cur, rens, '曲靖')
write_zhong_zhi(wb, cur, rens, '文山')
write_zhong_zhi(wb, cur, rens, '大理')
write_zhong_zhi(wb, cur, rens, '保山')
write_zhong_zhi(wb, cur, rens, '版纳')
write_zhong_zhi(wb, cur, rens, '怒江')
write_zhong_zhi(wb, cur, rens, '昭通')

rens = rens_list(cur, '后线')
write(wb, rens, '整体', '后线')
wb.move_sheet('后线人员整体保费排名', offset=-8)

wb.remove(wb['Sheet'])
wb.save("2019年销售人员业务跟踪表.xlsx")

cur.close()
conn.close()
