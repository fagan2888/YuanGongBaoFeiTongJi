import sqlite3
from openpyxl import Workbook
from stats import Stats


def rens_list(lei_xing):
    '''
    从数据库中获取前线人员和后线人员的名单
    并初始化人员信息列表对象
    '''

    year_sql = "SELECT MAX ([年份]) \
                FROM [销售人员业务跟踪表]"

    cur.execute(year_sql)
    max_year = cur.fetchone()[0]

    if lei_xing == '前线':
        str_sql = "SELECT 中支公司, 营销服务部, 业务员, 入司时间, 职级 FROM 前线"
    else:
        str_sql = "SELECT 中支公司, 营销服务部, 业务员, 入司时间 FROM 后线"

    cur.execute(str_sql)

    rens = []

    for values in cur.fetchall():
        info = Stats(cur)
        info.zhong_zhi = values[0][7:]
        info.ji_gou = values[1][11:]
        info.name = values[2][10:]
        info.ru_si_shi_jian = values[3][:10]
        if lei_xing == '前线':
            info.zhi_ji = values[4][16:]

        info.set_cai(nian_fen=f"{max_year}")
        info.set_che(nian_fen=f"{max_year}")
        info.set_ren(nian_fen=f"{max_year}")

        rens.append(info)

    return rens


def write(rens, xian_zhong, lei_xing='前线'):
    '''
    将分公司全体人员数据按险种写入到Excel中
    按保费规模进行排名
    '''
    if xian_zhong == '整体':
        rens_sort = sorted(rens, key=lambda ren: ren.zheng_ti, reverse=True)
    elif xian_zhong == '车险':
        rens_sort = sorted(rens, key=lambda ren: ren.che, reverse=True)
    else:
        rens_sort = sorted(rens, key=lambda ren: ren.fei_che, reverse=True)

    ws = wb.create_sheet(title=f'{lei_xing}人员{xian_zhong}保费排名')

    row = [f'{lei_xing}人员{xian_zhong}保费排名']
    ws.append(row)

    if lei_xing == '前线':
        row = ['排名',
               '姓名',
               f'{xian_zhong}保费（万元）',
               '机构',
               '职级',
               '入司时间']
    else:
        row = ['排名',
               '姓名',
               f'{xian_zhong}保费（万元）',
               '机构',
               '入司时间']
    ws.append(row)

    i = 1

    for ren in rens_sort:
        row = [i, ren.name]
        if xian_zhong == '整体':
            row.append(ren.zheng_ti)
        elif xian_zhong == '车险':
            row.append(ren.che)
        else:
            row.append(ren.fei_che)
        row.append(ren.ji_gou)
        if lei_xing == '前线':
            row.append(ren.zhi_ji)
        row.append(ren.ru_si_shi_jian)
        ws.append(row)
        i += 1


def write_zhong_zhi(rens, zhong_zhi):
    '''
    按中支为单位，将前线人员信息写入到Excel中
    以个人最新一年整体保费规模排名
    '''

    # 筛选出相应中支的人员
    new_rens = []
    for ren in rens:
        if zhong_zhi in ren.zhong_zhi:
            new_rens.append(ren)

    # 以整体保费规模进行排序
    rens_sort = sorted(new_rens, key=lambda ren: ren.zheng_ti, reverse=True)

    ws = wb.create_sheet(title=f'{zhong_zhi}前线人员整体保费跟踪表')

    row = [f'{zhong_zhi}前线人员整体保费跟踪表']
    ws.append(row)

    row = ['排名',
           '姓名',
           '机构',
           '职级',
           '入司时间']

    year_sql = "SELECT [年份] \
                FROM [销售人员业务跟踪表] \
                GROUP  BY [年份] \
                ORDER  BY [年份] DESC"

    cur.execute(year_sql)
    year = []
    for y_tupe in cur.fetchall():
        for y in y_tupe:
            year.append(y)

    for y in year:
        row.append(f'{y}年\n保费（万元）')

    for y in year:
        month_sql = f"SELECT [月份] \
                     FROM [销售人员业务跟踪表] \
                     WHERE [年份] = '{y}' \
                     GROUP  BY [月份] \
                     ORDER  BY [月份] DESC"

        cur.execute(month_sql)
        for m_tupe in cur.fetchall():
            for m in m_tupe:
                row.append(f"{y}年{m}月\n保费（万元）")

    ws.append(row)

    i = 1

    for ren in rens_sort:
        row = [i,
               ren.name,
               ren.ji_gou,
               ren.zhi_ji,
               ren.ru_si_shi_jian]
        for y in year:
            row.append(ren.bao_fei(nian_fen=y))

        for y in year:
            month_sql = f"SELECT [月份] \
                        FROM [销售人员业务跟踪表] \
                        WHERE [年份] = '{y}' \
                        GROUP  BY [月份] \
                        ORDER  BY [月份] DESC"

            cur.execute(month_sql)
            for m_tupe in cur.fetchall():
                for m in m_tupe:
                    row.append(ren.bao_fei(nian_fen=y, yue_fen=m))

        ws.append(row)
        i += 1


conn = sqlite3.connect('data.db')
cur = conn.cursor()

wb = Workbook()

rens = rens_list('前线')

# write(rens, '整体')
# write(rens, '车险')
# write(rens, '非车险')

write_zhong_zhi(rens, '分公司营业一部')
write_zhong_zhi(rens, '曲靖')
write_zhong_zhi(rens, '文山')
write_zhong_zhi(rens, '大理')
write_zhong_zhi(rens, '保山')
write_zhong_zhi(rens, '版纳')
write_zhong_zhi(rens, '怒江')
write_zhong_zhi(rens, '昭通')

# rens = rens_list('后线')
# write(rens, '整体', '后线')
# wb.move_sheet('后线人员整体保费排名', offset=-1)

wb.remove(wb['Sheet'])
wb.save("2019年销售人员业务跟踪表.xlsx")

cur.close()
conn.close()
